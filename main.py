from tkinter import*
from tkinter import ttk
import datetime
import openpyxl
from openpyxl import Workbook

root = Tk()  # class in tkinter
root.geometry('950x552')
#root.iconbitmap('D:\FCI\Dyanamic Languages\Project\Image\_123036.ico')
root.title('Market')

now = datetime.datetime.now()
date = now.strftime("%Y-%m-%d")

# =============Data Excel#=============
wb = Workbook()
ws = wb.active

ws.title = 'customer'

ws["A1"] = 'Full Name'
ws["B1"] = 'Number Phone'
ws["C1"] = 'Address'
ws["D1"] = 'Total'
ws["E1"] = 'Date Buy'

wb.save('customer.xlsx')
#-------------------------------------------

def save():
    name = En_Name.get()
    phone = En_Phone.get()
    Address = En_Address.get()
    total = En_Total.get()
    date = En_Date.get()

    excel = openpyxl.load_workbook('customer.xlsx')
    file = excel.active
    file.cell(column=1, row=file.max_row + 1, value=name)
    file.cell(column=2, row=file.max_row, value=phone)
    file.cell(column=3, row=file.max_row, value=Address)
    file.cell(column=4, row=file.max_row, value=total)
    file.cell(column=5, row=file.max_row, value=date)
    excel.save('customer.xlsx')


# =============Price#=============
menu = {
    0: ['GalaxyA50', 10000],
    1: ['GooglePixal', 8000],
    2: ['HuaweiMate20', 11000],
    3: ['Huawei P20', 13000],
    4: ['Honor', 6000],
    5: ['iphone 13 pro', 250000],
    6: ['iphone 14 pro max', 40000],
    7: ['iphone 8 plus', 11000],
}


def calc():
    global En_Name, En_Address, En_Date, En_Phone, En_Total
    lb_image.place(x=950,y=438, width=250 , height=110)
    root.geometry('1200x552')
    F4 = Frame(root, bg='#5F7161', width=250, height=434, bd=2)
    F4.place(x=950, y=1)

    L_Name = Label(F4, text='Name', bg='#5F7161', fg='white')
    L_Name.place(x=20, y=10)
    En_Name = Entry(F4, width=24, font=('Tajawal 12'), justify=CENTER)
    En_Name.place(x=15, y=40)

    L_Phone = Label(F4, text='Phone', bg='#5F7161', fg='white')
    L_Phone.place(x=20, y=70)
    En_Phone = Entry(F4, width=24, font=('Tajawal 12'), justify=CENTER)
    En_Phone.place(x=15, y=100)

    L_Address = Label(F4, text='Address', bg='#5F7161', fg='white')
    L_Address.place(x=20, y=130)
    En_Address = Entry(F4, width=24, font=('Tajawal 12'), justify=CENTER)
    En_Address.place(x=15, y=160)

    L_Total = Label(F4, text='Total', bg='#5F7161', fg='white')
    L_Total.place(x=20, y=190)
    En_Total = Entry(F4, width=24, font=('Tajawal 12'), justify=CENTER)
    En_Total.place(x=15, y=210)

    L_Date = Label(F4, text='Date', bg='#5F7161', fg='white')
    L_Date.place(x=20, y=240)
    En_Date = Entry(F4, width=24, font=('Tajawal 12'), justify=CENTER)
    En_Date.place(x=15, y=270)

    add_button = Button(F4, text='Save Receipt', width=31, cursor='hand2', bg='#EDDBC0', command=save)
    add_button.place(x=12, y=310)

    add_button = Button(F4, text='Empty fields', width=31, cursor='hand2', bg='#EDDBC0', command=clear1)
    add_button.place(x=12, y=340)

    total = 0
    for item in trv.get_children():  # delete all Things
        trv.delete(item)

    for i in range(len(sb)):
        if (int(sb[i].get()) > 0):
            price = int(sb[i].get()) * menu[i][1]
            total += price
            myst = (str(menu[i][1]), str(sb[i].get()), str(price)) #10000 , 2 , 20000
            trv.insert("", "end", iid=i, text=menu[i][0], values=myst)
            # GalaxyA50 ,10000 , 2 , 20000
    final = total
    En_Total.insert('1', str(final) + 'EGP')
    En_Date.insert('1', str(date))


def clear():
    for i in trv.get_children():
        trv.delete(i)
        
    var1 = IntVar() 
    var2 = IntVar()
    var3 = IntVar() 
    var4 = IntVar()
    var5 = IntVar() 
    var6 = IntVar()
    var7 = IntVar() 
    var8 = IntVar()
    var1.set(0),var2.set(0),var3.set(0),var4.set(0),var5.set(0),var6.set(0),var7.set(0),var7.set(0)
    sb1.config(textvariable=var1), sb2.config(textvariable=var2),sb3.config(textvariable=var3), sb4.config(textvariable=var4)
    sb5.config(textvariable=var5), sb6.config(textvariable=var6),sb7.config(textvariable=var7), sb8.config(textvariable=var8)
    En_Phone.delete('0', END)
    En_Address.delete('0', END)
    En_Date.delete('0', END)
    En_Total.delete('0', END)
    En_Name.delete('0',END)


def clear1():
    En_Name.delete('0', END)
    En_Phone.delete('0', END)
    En_Address.delete('0', END)
    En_Date.delete('0', END)
    En_Total.delete('0', END)


##=============Fram1#=============
F1 = Frame(root, bg='silver' , width=600 , height=800)
F1.place(x=1, y=1)

#=============Image#============= 
img_menu1 = PhotoImage(file='D:\FCI\Dyanamic Languages\Project\Image\galaxyA50_2.png')
img_menu2 = PhotoImage(file='D:\FCI\Dyanamic Languages\Project\Image\googlePixal_2.png')
img_menu3 = PhotoImage(file='D:\FCI\Dyanamic Languages\Project\Image\huwaieMate20.png')
img_menu4 = PhotoImage(file='D:\FCI\Dyanamic Languages\Project\Image\Huawei P20_2.png')
img_menu5 = PhotoImage(file='D:\FCI\Dyanamic Languages\Project\Image\Honor.png')
img_menu6 = PhotoImage(file='D:\FCI\Dyanamic Languages\Project\Image\iphone 13 pro_2.png')
img_menu7 = PhotoImage(file='D:\FCI\Dyanamic Languages\Project\Image\iphone 14 pro max_2.png')
img_menu8 = PhotoImage(file='D:\FCI\Dyanamic Languages\Project\Image\iphone 8 plus.png')


title = Label(F1, text='Mobile Store' , font=('Tajawal 13') ,fg='white' , bg='#5F7161',width=70)
title.place(x=0,y=0)
#7--45
menu1 = Button(F1 ,width=70,bg='#C0C0C0' , bd=1 , relief=FLAT ,cursor='hand2' , height=170,image=img_menu1 ,text= 'GalaxyA50' ,compound=TOP)
menu1.place(x=30,y=45)

menu2 = Button(F1 ,width=70,bg='#C0C0C0' , bd=1 , relief=FLAT ,cursor='hand2' , height=170,image=img_menu2 ,text= 'GooglePixal' ,compound=TOP)
menu2.place(x=170,y=45)

menu3 = Button(F1 ,width=73,bg='#C0C0C0' , bd=1 , relief=FLAT ,cursor='hand2' , height=170,image=img_menu3 ,text= 'HuaweiMate20' ,compound=TOP)
menu3.place(x=310,y=45) 

menu4 = Button(F1 ,width=70,bg='#C0C0C0' , bd=1 , relief=FLAT ,cursor='hand2' , height=170,image=img_menu4 ,text= 'Huawei P20' ,compound=TOP)
menu4.place(x=450,y=45) 

menu5 = Button(F1 ,width=70,bg='#C0C0C0' , bd=1 , relief=FLAT ,cursor='hand2' , height=170,image=img_menu5 ,text= 'Honor' ,compound=TOP)
menu5.place(x=30,y=267)

menu6 = Button(F1 ,width=70,bg='#C0C0C0' , bd=1 , relief=FLAT ,cursor='hand2' , height=170,image=img_menu6 ,text= 'Iphone 13 pro' ,compound=TOP)
menu6.place(x=170,y=267)

menu7 = Button(F1 ,width=73,bg='#C0C0C0' , bd=1 , relief=FLAT ,cursor='hand2' , height=170,image=img_menu7 ,text= 'Iphone14 pro' ,compound=TOP)
menu7.place(x=310,y=267)

menu8 = Button(F1 ,width=73,bg='#C0C0C0' , bd=1 , relief=FLAT ,cursor='hand2' , height=170,image=img_menu8 ,text= 'Iphone8 plus' ,compound=TOP)
menu8.place(x=450,y=267)


#=============Buttons=============
b1 = Button(F1,text='Buy' , fg='white' ,font=('Tajawal 12') ,width=15 ,bg='#6D8B74' ,bd=1,relief=SOLID, cursor='hand2' ,height=1 ,command=calc)
b1.place(x=60,y=500)

b2 = Button(F1,text='New Receipt' , fg='white' ,font=('Tajawal 12') ,width=15 ,bg='#6D8B74' ,bd=1,relief=SOLID, cursor='hand2' ,height=1 , command=clear)
b2.place(x=190,y=500)

b3 = Button(F1,text='Close' , fg='white' ,font=('Tajawal 12') ,width=15 ,bg='#6D8B74' ,bd=1,relief=SOLID, cursor='hand2' ,height=1,command=exit)
b3.place(x=320,y=500)



#=============Var and count =============
sb = []
font1 = ('Times' , 12 , 'normal')
sv1 =IntVar()
sv2 =IntVar()
sv3 =IntVar()
sv4 =IntVar()
sv5 =IntVar()
sv6 =IntVar()
sv7 =IntVar() 
sv8 =IntVar()

sb1 = Spinbox(F1, from_=0 ,to_=5, font=font1,justify="center", width=8, textvariable=sv1)
sb1.place(x=30,y=222)
sb.append(sb1)

sb2 = Spinbox(F1, from_=0 ,to_=5, font=font1, width=8,justify="center", textvariable=sv2)
sb2.place(x=170,y=222)
sb.append(sb2)

sb3 = Spinbox(F1, from_=0 ,to_=5, font=font1, width=8,justify="center", textvariable=sv3)
sb3.place(x=310,y=222)
sb.append(sb3)

sb4 = Spinbox(F1, from_=0 ,to_=5, font=font1, width=8,justify="center", textvariable=sv4)
sb4.place(x=450,y=222)
sb.append(sb4)

sb5 = Spinbox(F1, from_=0 ,to_=5, font=font1, width=8,justify="center", textvariable=sv5)
sb5.place(x=30,y=444)
sb.append(sb5)

sb6 = Spinbox(F1, from_=0 ,to_=5, font=font1, width=8,justify="center", textvariable=sv6)
sb6.place(x=170,y=444)
sb.append(sb6)

sb7 = Spinbox(F1, from_=0 ,to_=5, font=font1, width=8,justify="center", textvariable=sv7)
sb7.place(x=310,y=444)
sb.append(sb7)

sb8 = Spinbox(F1, from_=0 ,to_=5, font=font1, width=8,justify="center", textvariable=sv8)
sb8.place(x=450,y=444)
sb.append(sb8)

#=============Frame2=============
F2 = Frame(root, bg = 'gray' , width=343, height=550)
F2.place(x=604,y=1)

trv =  ttk.Treeview(F2,selectmode='browse')
trv.place(x=1,y=1,width=340, height=550)

trv["columns"] =('1','2','3')
trv.column("0" ,width=50 , anchor='c')
trv.column("1" ,width=50 , anchor='c')
trv.column("2" ,width=50 , anchor='c')
trv.column("3" ,width=50 , anchor='c')

trv.heading("#0" ,text='Mobile' ,anchor='c')
trv.heading("#1" ,text='Price' ,anchor='c')
trv.heading("#2" ,text='Num' ,anchor='c')
trv.heading("#3" ,text='Total' ,anchor='c')

im_logo = PhotoImage(file='D:\FCI\Dyanamic Languages\Project\Image\Mobile store logo.png')
lb_image= Label(root,image=im_logo)


root.mainloop()
